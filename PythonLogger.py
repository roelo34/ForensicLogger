import sys
import os.path
import os
import glob
import openpyxl
import datetime
import subprocess
import hashlib

def main():
    try:
        if sys.argv[1] == '-n':
            location = input('Please specify your case directory: ')
            case = input('Case name: ')
            researcher = input('Researcher name: ')
            researcherLocation = input('Location: ')
            screenshotsfolder = input('Screenshot folder: ')
            print('Setting up your log file...')
            logFile = openpyxl.Workbook()
            infoSheet = logFile.active
            infoSheet.title = 'Case Info'
            info = (
                ['Case', case],
                ['Location', researcherLocation],
                ['Researcher', researcher],
                ['Screenshot folder', screenshotsfolder]
            )
            for k in info:
                infoSheet.append(k)
            logFile.create_sheet(title='Log')
            logFile.create_sheet(title='CoC')
            titleFont = openpyxl.styles.Font(bold=True)
            logFile['Log'].append(['Who', 'When', 'Where', 'What', 'Why', 'How', 'Result'])
            logFile['CoC'].append(['Who', 'When', 'Where', 'File name', 'EID', 'Source', 'Destination', 'SHA1-Hash'])
            for cell in logFile['Log']['1:1']:
                cell.font = titleFont
            for cell in logFile['CoC']['1:1']:
                cell.font = titleFont
            logFile.save(filename = location + '/log.xlsx')
            print('Done')
        if sys.argv[1] == '-c':
            logFilePath = sys.argv[2] + '/log.xlsx'
            if glob.glob(logFilePath):
                logger(logFilePath)
            else:
                print('File does not exist!')
    except IndexError:
        print('Options:\nNew case\t-n\nExisting case\t-c <path to case>')

def logger(logFilePath):
    logFile = openpyxl.load_workbook(filename= logFilePath)
    name = logFile['Case Info']['B3'].value
    case = logFile['Case Info']['B1'].value
    screenshotsfolder = logFile['Case Info']['B4'].value
    storedLocations = logFile['Case Info']['B2'].value
    activeLocation = storedLocations.split(',')[-1]
    while True:
        command = input('\033[1m\033[92m' + case + ' $ ' + '\033[0m')
        if command == 'STOP':
            break
        elif command.split(' ')[0] == 'nano' or command.split(' ')[0] == 'vim' or command.split(' ')[0] == 'crontab' or command.split(' ')[0] == 'vi':
            print(command.split(' ')[0] + ' does not work properly with this app! Try using cat with grep instead.')
            continue
        elif command == 'CASEINFO':
            print('Case info\nResearcher:\t' + name + '\nCase name:\t' + case + '\nActive location:\t' + activeLocation)
            continue
        elif command == 'HELP' or command == '?':
            print('Help menu\nStop logging and save changes to log:\tSTOP\nGet info about the current case:\tCASEINFO\nManually log an action:\tMANUAL\nChange researcher location:\tCHANGELOC\nAdd record to CoC:\tCOC <file>')
        elif command == 'MANUAL':
            how = input('What action did you take? ')
            what = input('What do you want to accomplish by running this command? ')
            why = input('Why do you want to run this command? ')
            result = input('What is the result of this action? ')
            logFile['Log'].append([name, datetime.datetime.now(), activeLocation, what, why, how, result])
            continue
        elif command == 'SCREENSHOT':
            print("Screenshots will be saved at: {0}".format(screenshotsfolder))
            filename = input("Filename for the screenshot: ")
            os.system("import -window root {0}.png".format(os.path.join(screenshotsfolder, filename)))
        elif command == 'CHANGELOC':
            print('Stored locations: ' + storedLocations)
            newLocation = input('New location: ')
            if newLocation not in storedLocations:
                location = storedLocations + ',' + newLocation
                logFile['Case Info']['B2'] = location
                activeLocation = newLocation
                print('New locations will show up in the menu after a restart of the app.')
                continue
            else:
                activeLocation = newLocation
                continue
        elif command.split(' ')[0] == 'COC':
            EID = input('EID for ' + command.split(' ')[1] + ' : ')
            source = input('Source of ' + command.split(' ')[1] + ' : ')
            destination = input('Destination of ' + command.split(' ')[1] + ' : ')
            BUF_SIZE = 65536
            sha1 = hashlib.sha1()
            try:
                with open(command.split(' ')[1], 'rb') as f:
                    data = f.read(BUF_SIZE)
                    if not data:
                        break
                    sha1.update(data)
            except FileNotFoundError:
                print('File not found!')
                continue
            print(command.split(' ')[1] + '\t' + sha1.hexdigest())
            logFile['CoC'].append([name, datetime.datetime.now(), activeLocation, command.split(' ')[1], EID, source, destination, sha1.hexdigest()])
        else:
            log = input('Do you want to log this command? [Y/n]')
            if log == "n":
                os.system(command + '> /tmp/tmp')
                outputString = open('/tmp/tmp', 'r').read()
                print(outputString)
                os.remove('/tmp/tmp')
                continue
            what = input('What do you want to accomplish by running this command? ')
            why = input('Why do you want to run this command? ')
            os.system(command + '> /tmp/tmp')
            outputString = open('/tmp/tmp', 'r').read()
            print(outputString)
            logFile['Log'].append([name, datetime.datetime.now(), activeLocation, what, why, command, outputString])
            os.remove('/tmp/tmp')
    logFile.save(filename= logFilePath)

if __name__ == "__main__":
    main()
